import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import io
import base64
from datetime import datetime
import os

# Konfigurasi halaman
st.set_page_config(
    page_title="Aplikasi Inventarisasi Barang",
    page_icon="📦",
    layout="wide"
)

class InventoryAppStreamlit:
    def __init__(self):
        self.filename = "DBR.xlsx"
        self.initialize_session_state()
    
    def initialize_session_state(self):
        """Inisialisasi session state untuk menyimpan data sementara"""
        if 'data_barang' not in st.session_state:
            st.session_state.data_barang = []
        if 'edit_mode' not in st.session_state:
            st.session_state.edit_mode = False
        if 'edit_index' not in st.session_state:
            st.session_state.edit_index = None
    
    def is_merged_cell(self, sheet, row, col):
        """Cek apakah sel tertentu adalah bagian dari merged cells"""
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                return True
        return False
    
    def get_next_available_row(self, sheet, start_row=20):
        """Cari baris kosong berikutnya yang tidak termasuk dalam merged cells"""
        row = start_row
        max_checks = 100
        
        for _ in range(max_checks):
            if (sheet.cell(row=row, column=1).value is None and 
                not self.is_merged_cell(sheet, row, 1)):
                return row
            row += 1
        return row
    
    def format_template(self, sheet):
        """Format template Excel"""
        try:
            # Hapus merged cells yang mengganggu area data
            merged_ranges_to_remove = []
            for merged_range in list(sheet.merged_cells.ranges):
                if merged_range.min_row >= 20:
                    merged_ranges_to_remove.append(merged_range)
            
            for merged_range in merged_ranges_to_remove:
                sheet.unmerge_cells(str(merged_range))
            
            # Format header template
            if sheet['A10'].value is None:
                sheet.merge_cells('A10:I12')
                sheet['A10'] = "DAFTAR BARANG RUANGAN (DBR)"
                sheet['A10'].alignment = Alignment(horizontal='center', vertical='center')
                sheet['A10'].font = Font(size=14, bold=True)
            
            if sheet['A14'].value is None:
                sheet['A14'] = "Kode UAKPB"
                sheet['C14'] = ": 138.05.0200.693453.000KD"
                sheet['F14'] = "Ruangan"
                sheet['G14'] = ":"
                
                sheet['A15'] = "Nama Unit UAKPB"
                sheet['C15'] = ": BBPPMPV Pertanian Cianjur"
            
            if sheet['A17'].value is None:
                sheet.merge_cells('A17:I17')
                sheet['A17'] = "Nama Barang                 Tanda Pengenal Barang                                                             Keterangan"
                
                sheet['A18'] = "No."
                sheet['B18'] = "Nomor Urut"
                sheet['C18'] = "Nama Barang"
                sheet['D18'] = "Merk/"
                sheet['E18'] = "Kode Barang"
                sheet['F18'] = "Tahun"
                sheet['G18'] = "Jumlah"
                sheet['H18'] = "Keterangan"
                
                sheet['A19'] = "Urut"
                sheet['B19'] = "Pendaftaran"
                sheet['D19'] = "Type"
                sheet['F19'] = "Perolehan"
                sheet['G19'] = "Barang"
            
            if sheet['A50'].value is None:
                sheet.merge_cells('A50:I50')
                sheet['A50'] = f"Cianjur, {datetime.now().strftime('%B %Y')}"
                sheet['A50'].alignment = Alignment(horizontal='center')
                
                sheet['A52'] = "Penanggung Jawab UAKPB;"
                sheet['F52'] = "Penanggung Jawab Ruangan;"
                
                sheet['A53'] = "Kuasa Pengguna Barang"
                
                sheet['A55'] = "Dr. Yusuf, S.T., M.T."
                sheet['F55'] = "NIP. 196704181989121001"
                
                sheet['A56'] = "NIP. 196307201990011001"
                
        except Exception as e:
            st.warning(f"Peringatan dalam format template: {e}")
    
    def load_existing_data(self):
        """Muat data yang sudah ada dari file Excel"""
        try:
            if os.path.exists(self.filename):
                workbook = openpyxl.load_workbook(self.filename)
                if "TEMPLATE" in workbook.sheetnames:
                    sheet = workbook["TEMPLATE"]
                    
                    data = []
                    row_num = 20
                    
                    for row in sheet.iter_rows(min_row=20, max_row=100, values_only=True):
                        if (row[0] is not None and str(row[0]).strip() and 
                            not self.is_merged_cell(sheet, row_num, 1)):
                            data.append({
                                'No': row[0] or "",
                                'No Urut Pendaftaran': row[1] or "",
                                'Nama Barang': row[2] or "",
                                'Merk/Type': row[3] or "",
                                'Kode Barang': row[4] or "",
                                'Tahun Perolehan': row[5] or "",
                                'Jumlah': row[6] or "",
                                'Keterangan': row[7] or "",
                                'Excel_Row': row_num  # Simpan nomor baris Excel untuk referensi
                            })
                        row_num += 1
                    
                    workbook.close()
                    return data
            return []
        except Exception as e:
            st.error(f"Error membaca file: {e}")
            return []
    
    def save_to_excel(self, data):
        """Simpan data ke file Excel"""
        try:
            if os.path.exists(self.filename):
                workbook = openpyxl.load_workbook(self.filename)
            else:
                workbook = openpyxl.Workbook()
                default_sheet = workbook.active
                workbook.remove(default_sheet)
                workbook.create_sheet("TEMPLATE")
            
            if "TEMPLATE" not in workbook.sheetnames:
                workbook.create_sheet("TEMPLATE")
            
            sheet = workbook["TEMPLATE"]
            
            # Format template jika diperlukan
            self.format_template(sheet)
            
            # Cari baris kosong
            row_num = self.get_next_available_row(sheet, 20)
            
            # Tulis data
            sheet.cell(row=row_num, column=1, value=row_num-19)
            sheet.cell(row=row_num, column=2, value=data['no_urut'])
            sheet.cell(row=row_num, column=3, value=data['nama_barang'])
            sheet.cell(row=row_num, column=4, value=data['merk_type'])
            sheet.cell(row=row_num, column=5, value=data['kode_barang'])
            sheet.cell(row=row_num, column=6, value=data['tahun_perolehan'])
            sheet.cell(row=row_num, column=7, value=data['jumlah'])
            sheet.cell(row=row_num, column=8, value=data['keterangan'])
            
            workbook.save(self.filename)
            workbook.close()
            return True
            
        except Exception as e:
            st.error(f"Error menyimpan data: {e}")
            return False
    
    def update_in_excel(self, data, excel_row):
        """Update data di file Excel berdasarkan nomor baris"""
        try:
            if os.path.exists(self.filename):
                workbook = openpyxl.load_workbook(self.filename)
                sheet = workbook["TEMPLATE"]
                
                # Update data di baris yang spesifik
                sheet.cell(row=excel_row, column=1, value=data['No'])
                sheet.cell(row=excel_row, column=2, value=data['No Urut Pendaftaran'])
                sheet.cell(row=excel_row, column=3, value=data['Nama Barang'])
                sheet.cell(row=excel_row, column=4, value=data['Merk/Type'])
                sheet.cell(row=excel_row, column=5, value=data['Kode Barang'])
                sheet.cell(row=excel_row, column=6, value=data['Tahun Perolehan'])
                sheet.cell(row=excel_row, column=7, value=data['Jumlah'])
                sheet.cell(row=excel_row, column=8, value=data['Keterangan'])
                
                workbook.save(self.filename)
                workbook.close()
                return True
            return False
        except Exception as e:
            st.error(f"Error mengupdate data: {e}")
            return False
    
    def delete_from_excel(self, excel_row):
        """Hapus data dari file Excel berdasarkan nomor baris"""
        try:
            if os.path.exists(self.filename):
                workbook = openpyxl.load_workbook(self.filename)
                sheet = workbook["TEMPLATE"]
                
                # Hapus baris dengan menggeser ke atas
                sheet.delete_rows(excel_row)
                
                workbook.save(self.filename)
                workbook.close()
                return True
            return False
        except Exception as e:
            st.error(f"Error menghapus data: {e}")
            return False
    
    def get_excel_download_link(self):
        """Generate link untuk download file Excel"""
        try:
            if os.path.exists(self.filename):
                with open(self.filename, "rb") as file:
                    excel_data = file.read()
                
                b64 = base64.b64encode(excel_data).decode()
                href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{self.filename}">📥 Download File Excel</a>'
                return href
            return None
        except Exception as e:
            st.error(f"Error generating download link: {e}")
            return None
    
    def render_form(self):
        """Render form input data"""
        st.header("📝 Form Input Data Barang")
        
        with st.form("input_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                no_urut = st.text_input("Nomor Urut Pendaftaran*", 
                                      placeholder="Masukkan nomor urut pendaftaran")
                nama_barang = st.text_input("Nama Barang*", 
                                          placeholder="Masukkan nama barang")
                merk_type = st.text_input("Merk/Type", 
                                        placeholder="Masukkan merk/type")
                kode_barang = st.text_input("Kode Barang", 
                                          placeholder="Masukkan kode barang")
            
            with col2:
                tahun_perolehan = st.text_input("Tahun Perolehan", 
                                              placeholder="Masukkan tahun perolehan")
                jumlah = st.text_input("Jumlah Barang", 
                                     placeholder="Masukkan jumlah barang")
                keterangan = st.selectbox("Keterangan", 
                                        ["Baik", "Rusak", "Usang", "Hibah Pusdatin", "Lainnya"])
            
            submitted = st.form_submit_button("💾 Simpan Data")
            
            if submitted:
                if not no_urut or not nama_barang:
                    st.error("❌ Nomor Urut Pendaftaran dan Nama Barang harus diisi!")
                else:
                    data = {
                        'no_urut': no_urut,
                        'nama_barang': nama_barang,
                        'merk_type': merk_type,
                        'kode_barang': kode_barang,
                        'tahun_perolehan': tahun_perolehan,
                        'jumlah': jumlah,
                        'keterangan': keterangan
                    }
                    
                    if self.save_to_excel(data):
                        st.success("✅ Data berhasil disimpan!")
                        st.session_state.data_barang = self.load_existing_data()
    
    def render_data_table(self):
        """Render tabel data barang dengan fitur edit dan hapus per baris"""
        st.header("📊 Data Barang Inventaris")
        
        # Muat data
        data = self.load_existing_data()
        
        if data:
            # Tampilkan data dalam format tabel dengan fitur edit/hapus
            for i, item in enumerate(data):
                with st.container():
                    col1, col2, col3, col4, col5, col6, col7, col8, col9, col10 = st.columns([1, 2, 3, 2, 2, 2, 1, 2, 1, 1])
                    
                    with col1:
                        st.write(f"**{item['No']}**")
                    
                    with col2:
                        no_urut = st.text_input(
                            "No Urut", 
                            value=item['No Urut Pendaftaran'],
                            key=f"no_urut_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col3:
                        nama_barang = st.text_input(
                            "Nama Barang",
                            value=item['Nama Barang'],
                            key=f"nama_barang_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col4:
                        merk_type = st.text_input(
                            "Merk/Type",
                            value=item['Merk/Type'],
                            key=f"merk_type_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col5:
                        kode_barang = st.text_input(
                            "Kode Barang",
                            value=item['Kode Barang'],
                            key=f"kode_barang_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col6:
                        tahun_perolehan = st.text_input(
                            "Tahun Perolehan",
                            value=item['Tahun Perolehan'],
                            key=f"tahun_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col7:
                        jumlah = st.text_input(
                            "Jumlah",
                            value=item['Jumlah'],
                            key=f"jumlah_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col8:
                        keterangan = st.selectbox(
                            "Keterangan",
                            ["Baik", "Rusak", "Usang", "Hibah Pusdatin", "Lainnya"],
                            index=["Baik", "Rusak", "Usang", "Hibah Pusdatin", "Lainnya"].index(item['Keterangan']) if item['Keterangan'] in ["Baik", "Rusak", "Usang", "Hibah Pusdatin", "Lainnya"] else 0,
                            key=f"keterangan_{i}",
                            label_visibility="collapsed"
                        )
                    
                    with col9:
                        if st.button("✏️", key=f"edit_{i}", help="Edit data"):
                            updated_data = {
                                'No': item['No'],
                                'No Urut Pendaftaran': no_urut,
                                'Nama Barang': nama_barang,
                                'Merk/Type': merk_type,
                                'Kode Barang': kode_barang,
                                'Tahun Perolehan': tahun_perolehan,
                                'Jumlah': jumlah,
                                'Keterangan': keterangan
                            }
                            
                            if self.update_in_excel(updated_data, item['Excel_Row']):
                                st.success(f"✅ Data {nama_barang} berhasil diupdate!")
                                st.session_state.data_barang = self.load_existing_data()
                                st.rerun()
                    
                    with col10:
                        if st.button("🗑️", key=f"delete_{i}", help="Hapus data"):
                            if self.delete_from_excel(item['Excel_Row']):
                                st.success(f"✅ Data {item['Nama Barang']} berhasil dihapus!")
                                st.session_state.data_barang = self.load_existing_data()
                                st.rerun()
                
                st.divider()
            
            # Tombol refresh data
            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("🔄 Refresh Data"):
                    st.session_state.data_barang = self.load_existing_data()
                    st.rerun()
            
            # Download link untuk file Excel lengkap
            st.markdown("---")
            st.subheader("📥 Download Data")
            download_link = self.get_excel_download_link()
            if download_link:
                st.markdown(download_link, unsafe_allow_html=True)
            else:
                st.info("File Excel belum tersedia untuk di-download")
                
        else:
            st.info("📭 Belum ada data barang. Silakan input data terlebih dahulu.")
    
    def run(self):
        """Jalankan aplikasi"""
        st.title("📦 Aplikasi Inventarisasi Barang")
        st.markdown("---")
        
        # Tab untuk organisasi yang lebih baik
        tab1, tab2, tab3 = st.tabs(["Input Data", "Data Barang", "Informasi"])
        
        with tab1:
            self.render_form()
        
        with tab2:
            self.render_data_table()
        
        with tab3:
            st.header("ℹ️ Informasi Aplikasi")
            st.markdown("""
            ### Cara Penggunaan:
            1. **Input Data**: Gunakan tab "Input Data" untuk menambahkan barang inventaris
            2. **Edit/Hapus Data**: Gunakan tab "Data Barang" untuk mengedit atau menghapus data langsung di setiap baris
            3. **Download**: Gunakan link download untuk mendapatkan file Excel lengkap
            
            ### Fitur Baru:
            - ✏️ **Edit Data**: Klik tombol edit (✏️) di setiap baris untuk mengupdate data
            - 🗑️ **Hapus Data**: Klik tombol hapus (🗑️) di setiap baris untuk menghapus data
            - 🔄 **Refresh Data**: Memperbarui tampilan data terbaru
            - 📥 **Download Excel**: Mendownload file Excel dengan format DBR
            
            ### Kolom Wajib:
            - ❗ **Nomor Urut Pendaftaran**
            - ❗ **Nama Barang**
            
            ### Keterangan Status:
            - ✅ **Baik**: Barang dalam kondisi baik
            - ❌ **Rusak**: Barang perlu perbaikan
            - ⚠️ **Usang**: Barang sudah tua tapi masih berfungsi
            - 🎁 **Hibah Pusdatin**: Barang dari hibah
            """)

# Jalankan aplikasi
if __name__ == "__main__":
    app = InventoryAppStreamlit()
    app.run()
